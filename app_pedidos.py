#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Leitura de Código de Barras - Versão com Filtro por Pedido
Com cores únicas por pedido
"""

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import io
from datetime import datetime

# Configuração da página
st.set_page_config(
    page_title="Sistema de Código de Barras - Pedidos",
    page_icon="📦",
    layout="wide"
)

# Cores disponíveis para destaque de produtos escaneados
CORES_DESTAQUE_SCAN = [
    "FFFF00",  # Amarelo
    "90EE90",  # Verde claro
    "FFB6C1",  # Rosa claro
    "87CEEB",  # Azul claro
    "FFD700",  # Dourado
    "FFA500",  # Laranja
    "DDA0DD",  # Ameixa
    "F0E68C",  # Cáqui
]

# Cores para pedidos (tons mais suaves)
CORES_PEDIDOS = [
    "E8F4F8",  # Azul muito claro
    "FFF4E6",  # Laranja muito claro
    "F0F8E8",  # Verde muito claro
    "FFF0F5",  # Rosa muito claro
    "F5F5DC",  # Bege
    "E6E6FA",  # Lavanda
    "F0FFF0",  # Verde menta claro
    "FFF8DC",  # Cornsilk
]

# Inicializar estado da sessão
if 'df' not in st.session_state:
    st.session_state.df = None
if 'df_filtrado' not in st.session_state:
    st.session_state.df_filtrado = None
if 'produtos_escaneados' not in st.session_state:
    st.session_state.produtos_escaneados = []
if 'indice_cor' not in st.session_state:
    st.session_state.indice_cor = 0
if 'ultima_linha_encontrada' not in st.session_state:
    st.session_state.ultima_linha_encontrada = None
if 'pedido_filtrado' not in st.session_state:
    st.session_state.pedido_filtrado = None
if 'cores_por_pedido' not in st.session_state:
    st.session_state.cores_por_pedido = {}


def exportar_excel_com_destaque(df, produtos_escaneados, cores_por_pedido, coluna_pedido):
    """Exporta planilha Excel com destaques aplicados"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Escrever cabeçalhos
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Escrever dados
    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Aplicar cores de pedido primeiro (fundo)
    if coluna_pedido in df.columns:
        for idx, row in df.iterrows():
            pedido = str(row[coluna_pedido])
            if pedido in cores_por_pedido:
                linha_excel = idx + 2
                cor = cores_por_pedido[pedido]
                fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
                
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=linha_excel, column=col).fill = fill
    
    # Aplicar destaques de produtos escaneados (sobrepõe)
    for produto in produtos_escaneados:
        linha_excel = produto['linha'] + 2
        cor = produto['cor']
        fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
        
        for col in range(1, ws.max_column + 1):
            ws.cell(row=linha_excel, column=col).fill = fill
    
    # Salvar em buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


st.title("📦 Sistema de Código de Barras - Gestão de Pedidos")

# Upload de arquivo
arquivo_upload = st.file_uploader("Selecione a planilha Excel", type=['xlsx', 'xls', 'xlsm'])

if arquivo_upload is not None:
    try:
        df_original = pd.read_excel(arquivo_upload, engine='openpyxl')
        st.session_state.df = df_original
        
        st.success(f"✓ Planilha carregada: {arquivo_upload.name}")
        st.info(f"📊 Total de linhas: {len(df_original)}")
        
        # Detectar coluna de código
        colunas_possiveis = ['EAN_PRODUTO', 'EAN', 'ean', 'Código de Barras', 'codigo', 'barcode']
        coluna_codigo = None
        
        for col in colunas_possiveis:
            if col in df_original.columns:
                coluna_codigo = col
                break
        
        if coluna_codigo:
            st.success(f"✓ Coluna de código detectada: **{coluna_codigo}**")
        else:
            st.warning("⚠️ Coluna de código não detectada automaticamente")
            coluna_codigo = st.selectbox("Selecione a coluna de código:", df_original.columns)
        
        # Detectar coluna de pedido
        coluna_pedido = 'PEDIDO' if 'PEDIDO' in df_original.columns else None
        
        if coluna_pedido:
            st.success(f"✓ Coluna de pedido detectada: **{coluna_pedido}**")
        else:
            st.warning("⚠️ Coluna PEDIDO não encontrada")
            coluna_pedido = st.selectbox("Selecione a coluna de pedido:", df_original.columns)
        
        st.divider()
        
        # FILTRO POR PEDIDO
        st.header("🔍 Filtrar por Pedido")
        
        col_filtro1, col_filtro2 = st.columns([4, 1])
        
        with col_filtro1:
            # Obter lista de pedidos únicos
            pedidos_unicos = sorted(df_original[coluna_pedido].unique())
            
            pedido_input = st.selectbox(
                "Selecione o número do pedido:",
                options=["Todos"] + [str(p) for p in pedidos_unicos],
                key="pedido_select"
            )
        
        with col_filtro2:
            st.markdown("<br>", unsafe_allow_html=True)
            aplicar_filtro = st.button("✅ Aplicar", type="primary", use_container_width=True)
        
        # Aplicar filtro
        if aplicar_filtro:
            if pedido_input == "Todos":
                st.session_state.df_filtrado = df_original.copy()
                st.session_state.pedido_filtrado = None
                st.info("📋 Mostrando todos os pedidos")
            else:
                st.session_state.df_filtrado = df_original[df_original[coluna_pedido] == int(pedido_input)].copy()
                st.session_state.pedido_filtrado = pedido_input
                st.success(f"✓ Filtrado para pedido: **{pedido_input}** ({len(st.session_state.df_filtrado)} produtos)")
                
                # Atribuir cor ao pedido se ainda não tiver
                if pedido_input not in st.session_state.cores_por_pedido:
                    cor_idx = len(st.session_state.cores_por_pedido) % len(CORES_PEDIDOS)
                    st.session_state.cores_por_pedido[pedido_input] = CORES_PEDIDOS[cor_idx]
        
        # Usar DataFrame filtrado ou original
        df = st.session_state.df_filtrado if st.session_state.df_filtrado is not None else df_original
        
        st.divider()
        
        #Estatísticas
        #col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
        #with col_stat1:
            #st.metric("Total de Produtos", len(df))
        #with col_stat2:
            #st.metric("Produtos Escaneados", len(st.session_state.produtos_escaneados))
        #with col_stat3:
            #pedidos_na_view = df[coluna_pedido].nunique()
            #st.metric("Pedidos na Visualização", pedidos_na_view)
        #with col_stat4:
            #if len(df) > 0:
                #progresso = (len(st.session_state.produtos_escaneados) / len(df)) * 100
                #st.metric("Progresso", f"{progresso:.1f}%")
        
        #st.divider()
        
        # Leitor de código
        st.header("🔍 Leitor de Código de Barras")
        
        col1, col2 = st.columns([5, 1])
        
        with col1:
            codigo_input = st.text_input(
                "Digite ou escaneie o código:",
                key="codigo_barras",
                placeholder="Posicione o cursor aqui e escaneie com o leitor USB..."
            )
        
        with col2:
            st.markdown("<br>", unsafe_allow_html=True)
            buscar = st.button("🔎 Buscar", type="primary", use_container_width=True)
        
        if codigo_input and buscar:
            st.write("---")
            
            codigo_busca = str(codigo_input).strip()
            
            # Buscar produto no DataFrame filtrado
            try:
                resultado = df[df[coluna_codigo] == int(codigo_busca)]
            except Exception:
                resultado = pd.DataFrame()
            
            if resultado.empty:
                try:
                    resultado = df[df[coluna_codigo].astype(str) == codigo_busca]
                except Exception:
                    resultado = pd.DataFrame()
            
            if resultado.empty:
                try:
                    df_temp = df.copy()
                    df_temp['_codigo_str'] = df_temp[coluna_codigo].astype(str).str.strip().str.replace('.0', '', regex=False)
                    resultado = df_temp[df_temp['_codigo_str'] == codigo_busca]
                except Exception:
                    resultado = pd.DataFrame()
            
            if not resultado.empty:
                linha_idx = resultado.index[0]
                
                # Guardar última linha encontrada
                st.session_state.ultima_linha_encontrada = linha_idx
                
                # Calcular posição na tabela filtrada
                posicao_tabela = df.index.get_loc(linha_idx) + 1
                total_linhas = len(df)
                
                st.success(f"✅ PRODUTO ENCONTRADO!")
                
                # Mostrar posição
                st.info(f"📍 **Posição na tabela: Linha {posicao_tabela} de {total_linhas}**")
                
                if len(resultado) > 1:
                    st.warning(f"⚠️ Encontrados {len(resultado)} produtos com este código. Mostrando o primeiro.")
                
                # Verificar se já foi escaneado
                ja_escaneado = any(p['codigo'] == codigo_input for p in st.session_state.produtos_escaneados)
                
                if not ja_escaneado:
                    # Adicionar aos escaneados
                    cor = CORES_DESTAQUE_SCAN[st.session_state.indice_cor % len(CORES_DESTAQUE_SCAN)]
                    st.session_state.indice_cor += 1
                    
                    produto_info = {
                        'codigo': codigo_input,
                        'linha': linha_idx,
                        'cor': cor,
                        'timestamp': datetime.now()
                    }
                    
                    st.session_state.produtos_escaneados.append(produto_info)
                    
                    #st.success(f"🎨 Produto destacado com a cor: #{cor}")
                else:
                    st.info("ℹ️ Este produto já foi escaneado anteriormente.")
                
                # Mostrar informações do produto
                st.write("### 📦 Informações do Produto")
                
                # Criar tabela de informações
                info_data = []
                for col_name, value in resultado.iloc[0].items():
                    info_data.append({"Campo": col_name, "Valor": value})
                
                st.dataframe(pd.DataFrame(info_data), use_container_width=True, hide_index=True)
                
            else:
                st.error(f"❌ Produto não encontrado: {codigo_input}")
                if st.session_state.pedido_filtrado:
                    st.warning(f"⚠️ Lembre-se: você está filtrando apenas o pedido {st.session_state.pedido_filtrado}")
        
        st.divider()
        
        # Botões de ação
        col_btn1, col_btn2, col_btn3 = st.columns(3)
        
        with col_btn1:
            if st.button("🔄 Limpar Destaques", use_container_width=True):
                st.session_state.produtos_escaneados = []
                st.session_state.indice_cor = 0
                st.session_state.ultima_linha_encontrada = None
                st.rerun()
        
        with col_btn2:
            if st.button("🗑️ Limpar Filtro", use_container_width=True):
                st.session_state.df_filtrado = None
                st.session_state.pedido_filtrado = None
                st.rerun()
        
        with col_btn3:
            if len(st.session_state.produtos_escaneados) > 0 or len(st.session_state.cores_por_pedido) > 0:
                buffer = exportar_excel_com_destaque(
                    df_original, 
                    st.session_state.produtos_escaneados,
                    st.session_state.cores_por_pedido,
                    coluna_pedido
                )
                
                st.download_button(
                    label="📥 Baixar Excel com Destaques",
                    data=buffer,
                    file_name=f"verificados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        
        st.divider()
        
        # Mostrar cores de pedidos
        #if len(st.session_state.cores_por_pedido) > 0:
            #st.header("🎨 Cores por Pedido")
            #cores_info = []
            #for pedido, cor in st.session_state.cores_por_pedido.items():
                #cores_info.append({
                    #"Pedido": pedido,
                    #"Cor": f"#{cor}",
                    #"Produtos": len(df_original[df_original[coluna_pedido] == int(pedido)])
                #})
            #st.dataframe(pd.DataFrame(cores_info), use_container_width=True, hide_index=True)
            #st.divider()
        
        # Mostrar tabela com destaques
        st.header("📊 Planilha com Destaques")
        
        if st.session_state.ultima_linha_encontrada is not None:
            posicao = df.index.get_loc(st.session_state.ultima_linha_encontrada) + 1 if st.session_state.ultima_linha_encontrada in df.index else None
            if posicao:
                st.info(f"👉 Role a tabela abaixo para ver a linha {posicao} destacada")
        
        st.write(f"**Produtos escaneados:** {len(st.session_state.produtos_escaneados)}")
        
        # Aplicar destaque
        def highlight_rows(row):
            # Primeiro verificar se é produto escaneado (prioridade)
            for produto in st.session_state.produtos_escaneados:
                if row.name == produto['linha']:
                    cor_hex = produto['cor']
                    r = int(cor_hex[0:2], 16)
                    g = int(cor_hex[2:4], 16)
                    b = int(cor_hex[4:6], 16)
                    return [f'background-color: rgb({r},{g},{b}); font-weight: bold'] * len(row)
            
            # Se não for escaneado, aplicar cor do pedido
            if coluna_pedido in row.index:
                pedido = str(row[coluna_pedido])
                if pedido in st.session_state.cores_por_pedido:
                    cor_hex = st.session_state.cores_por_pedido[pedido]
                    r = int(cor_hex[0:2], 16)
                    g = int(cor_hex[2:4], 16)
                    b = int(cor_hex[4:6], 16)
                    return [f'background-color: rgb({r},{g},{b})'] * len(row)
            
            return [''] * len(row)
        
        # Criar DataFrame para exibição com indicador de linha
        df_display = df.copy()
        df_display.insert(0, '📍 Linha', range(1, len(df_display) + 1))
        
        st.dataframe(
            df_display.style.apply(highlight_rows, axis=1),
            use_container_width=True,
            height=500
        )
        
        # Mostrar produtos escaneados
        if len(st.session_state.produtos_escaneados) > 0:
            st.divider()
            st.header("✅ Produtos Escaneados")
            
            escaneados_data = []
            for p in st.session_state.produtos_escaneados:
                # Buscar informações do produto
                produto_row = df_original.loc[p['linha']]
                escaneados_data.append({
                    "Linha": p['linha'] + 1,
                    "Pedido": produto_row[coluna_pedido],
                    "Código": p['codigo'],
                    "Produto": produto_row.get('DESCRICAO_MODELO', ''),
                    "Cor Destaque": f"#{p['cor']}",
                    "Horário": p['timestamp'].strftime("%H:%M:%S")
                })
            
            st.dataframe(pd.DataFrame(escaneados_data), use_container_width=True, hide_index=True)
        
    except Exception as e:
        st.error(f"Erro ao carregar planilha: {e}")
        import traceback
        st.code(traceback.format_exc())
else:
    st.info("👆 Faça upload de uma planilha Excel para começar")
    
    st.markdown("---")
    st.markdown("### 📋 Como Usar")
    st.markdown("""
    1. **Carregue sua planilha** Excel (xlsx, xls ou xlsm)
    2. **Filtre por pedido** (opcional) para ver apenas produtos de um pedido específico
    3. **Escaneie produtos** com o leitor USB
    4. **Visualize destaques:**
       - Cor suave de fundo = Pedido
       - Cor forte = Produto escaneado
    5. **Baixe o Excel** com todos os destaques aplicados
    """)

# Footer
st.divider()
st.markdown(
    "<div style='text-align: center; color: gray;'>Sistema de Código de Barras - Gestão de Pedidos | Desenvolvido por Fabio</div>",
    unsafe_allow_html=True
)

